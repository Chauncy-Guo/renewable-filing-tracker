# -*- coding: utf-8 -*-
"""通用 Excel 写入工具。"""
from typing import List, Sequence
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def save_xlsx(rows: Sequence[Sequence], out_path: str, sheet_name: str = "数据") -> str:
    """把二维表格保存为带表头样式的 xlsx 文件。

    Args:
        rows: 二维列表, 第一行必须是表头。
        out_path: 输出 .xlsx 路径。
        sheet_name: Sheet 名 (最长 31 字符)。

    Returns:
        实际写入的路径。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, row in enumerate(rows, 1):
        if not row or all(not str(c).strip() for c in row):
            continue
        for ci, val in enumerate(row, 1):
            try:
                cell = ws.cell(row=ri, column=ci, value=val)
            except Exception:
                cell = ws.cell(row=ri, column=ci, value=str(val)[:1000])
            cell.alignment = Alignment(
                horizontal="left" if ci > 1 else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
            if ri == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

    # 列宽 (中文字符按 2 算)
    for col in ws.columns:
        max_len = 8
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = str(cell.value) if cell.value else ""
            for ch in v:
                max_len += 2 if ord(ch) > 127 else 1
            max_len = min(max_len, 50)
        ws.column_dimensions[col_letter].width = max(12, min(max_len, 50))

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path
