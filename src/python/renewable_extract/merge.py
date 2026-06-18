# -*- coding: utf-8 -*-
"""把 Excel/ 中的单月文件按类型汇总到 汇总/。"""
import os
import re
from typing import List
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .config import XLSX_DIR, OUT_DIR, HEADER

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
CENTER = Alignment(horizontal="center", vertical="center")
MONTH_RE = re.compile(r"(?:^|[^0-9])(\d{2,4})年(\d{1,2})月")


def parse_month(filename: str):
    """从文件名提取 (年, 月), 年份归一为 4 位。失败返回 (0, 0)。"""
    m = MONTH_RE.search(filename)
    if not m:
        return (0, 0)
    y = int(m.group(1))
    if y < 100:
        y = 2000 + y
    return (y, int(m.group(2)))


def fmt_month(year: int, month: int) -> str:
    return f"{year}年{month}月份"


def _load_xlsx_rows(filepath: str):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _merge_category(cat_name: str, file_predicate, with_month_col: bool):
    files = []
    for f in os.listdir(XLSX_DIR):
        if not f.endswith(".xlsx"):
            continue
        if not file_predicate(f):
            continue
        ym = parse_month(f)
        if ym == (0, 0):
            continue
        files.append((ym, f))
    if not files:
        print(f"  [跳过] {cat_name}: 无文件")
        return None

    files.sort(key=lambda x: x[0], reverse=True)
    print(f"  {cat_name}: {len(files)} 个文件")
    for ym, f in files:
        print(f"    {fmt_month(*ym)} <- {f}")

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = cat_name
    header = HEADER if with_month_col else HEADER[:5]
    ws.append(header)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for ym, f in files:
        year, month = ym
        rows = _load_xlsx_rows(os.path.join(XLSX_DIR, f))
        if not rows:
            continue
        data_rows = rows[1:] if rows[0] and rows[0][0] == "序号" else rows
        month_str = fmt_month(year, month)
        for r in data_rows:
            if r is None or len(r) == 0:
                continue
            if not any(c for c in r[:4] if c):
                continue
            row = list(r[:5])
            while len(row) < 5:
                row.append(None)
            if with_month_col:
                row.append(month_str)
            ws.append(row)

    widths = [8, 22, 50, 12, 30, 14]
    for i, w in enumerate(widths[:len(header)]):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    out_name = f"汇总_{cat_name}.xlsx"
    out_path = os.path.join(SUMMARY_DIR, out_name)
    out_wb.save(out_path)
    print(f"    -> 已保存: {out_path}  ({ws.max_row - 1} 条数据)")
    return out_path


# 过滤谓词
def is_industrial(f: str) -> bool:
    return "工商业分布式" in f


def is_central(f: str) -> bool:
    return "集中式光伏" in f


def is_wind(f: str) -> bool:
    return "风机" in f


def is_biomass(f: str) -> bool:
    return "生物质" in f


def merge_all(verbose: bool = True) -> List[str]:
    """汇总所有 4 类: 工商业分布式 / 集中式光伏 / 风机 / 生物质。"""
    if verbose:
        print("=" * 60)
        print("开始生成汇总 Excel")
        print("=" * 60)
    paths = []
    for name, pred, with_col in [
        ("工商业分布式光伏", is_industrial, True),
        ("集中式光伏", is_central, True),
        ("风机", is_wind, True),
        ("生物质", is_biomass, True),
    ]:
        p = _merge_category(name, pred, with_col)
        if p:
            paths.append(p)
    if verbose:
        print("\n完成!")
    return paths
